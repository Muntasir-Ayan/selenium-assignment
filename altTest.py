from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook

# Function to test for missing 'alt' attribute in images
def test_image_alt_attribute(url, testcase):
    # List to store results for each image
    image_results = []
    
    # Initialize the WebDriver
    driver = webdriver.Chrome()

    try:
        # Open the URL
        driver.get(url)

        # Find all <img> elements
        images = driver.find_elements(By.TAG_NAME, "img")
        
        # Check if each image has the 'alt' attribute
        for img in images:
            image_src = img.get_attribute("src")
            alt_attr = img.get_attribute("alt")
            
            if alt_attr:
                result = {
                    "page_url": url,
                    "testcase": testcase,
                    "image_src": image_src,
                    "passed/fail": "passed",
                    "comments": f"Alt attribute present: {alt_attr}"
                }
            else:
                result = {
                    "page_url": url,
                    "testcase": testcase,
                    "image_src": image_src,
                    "passed/fail": "fail",
                    "comments": "Missing 'alt' attribute"
                }
            
            # Add the result for this image to the list
            image_results.append(result)

    except Exception as e:
        result = {
            "page_url": url,
            "testcase": testcase,
            "image_src": "N/A",
            "passed/fail": "fail",
            "comments": str(e)
        }
        image_results.append(result)
    
    finally:
        # Close the driver after use
        driver.quit()

    return image_results

# Function to add results to the Excel file
def add_to_excel(test_results, excel_file):
    try:
        # Load the existing workbook
        workbook = load_workbook(excel_file)
        
        # Check if the sheet already exists
        if "Image_Alt_Attribute_Test" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Image_Alt_Attribute_Test")
        else:
            worksheet = workbook["Image_Alt_Attribute_Test"]
        
        # If the sheet is empty, add headers
        if worksheet.max_row == 1:
            worksheet.append(["page_url", "testcase", "image_src", "passed/fail", "comments"])
        
        # Add each result to the sheet
        for result in test_results:
            worksheet.append([result["page_url"], result["testcase"], result["image_src"], result["passed/fail"], result["comments"]])
        
        # Save the workbook with the new sheet data
        workbook.save(excel_file)
        print("Test results added to the 'Image_Alt_Attribute_Test' sheet successfully!")

    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")

# List of test cases (URLs and test case names)
test_cases = [
    {"url": "https://www.alojamiento.io/property/campo-lindo-apartment/BC-1935047", "testcase": "Image Alt Attribute Test"},
    # Add more test cases here if needed
]

# List to hold test results
test_results = []

# Run the tests for each case
for case in test_cases:
    result = test_image_alt_attribute(case["url"], case["testcase"])
    test_results.extend(result)  # Add each image result to the final results

# Add the results to an existing Excel file
add_to_excel(test_results, "test_report.xlsx")



print("alt Test report generated successfully!")