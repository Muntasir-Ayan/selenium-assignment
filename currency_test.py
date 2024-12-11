from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pandas as pd
from openpyxl import load_workbook

# List to store results
results = []

# Define the default currency (EUR in this case)
default_currency = "€"

try:
    # Initialize WebDriver
    driver = webdriver.Chrome()

    # Navigate to the URL
    url = "https://www.alojamiento.io/property/campo-lindo-apartment/BC-1935047"
    driver.get(url)
    driver.maximize_window()

    # Wait for the page to fully load
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "js-currency-sort-footer")))

    # Locate the currency dropdown
    currency_dropdown = driver.find_element(By.ID, "js-currency-sort-footer")
    action = ActionChains(driver)
    action.move_to_element(currency_dropdown).perform()  # Ensure dropdown is visible

    # Click the dropdown to expand it
    driver.execute_script("arguments[0].click();", currency_dropdown)
    time.sleep(1)  # Allow time for the dropdown to expand

    # Fetch <li> elements in the dropdown
    li_elements = driver.find_elements(By.XPATH, "//*[@id='js-currency-sort-footer']//ul[@class='select-ul']//li")

    print(f"Number of <li> elements: {len(li_elements)}")

    # Fetch initial prices
    price_elements = driver.find_elements(By.CLASS_NAME, 'js-price-value')
    initial_prices = [elem.text for elem in price_elements]
    print(f"Found {len(initial_prices)} initial price elements.")

    # Iterate and click each currency option
    for i in range(len(li_elements)):
        # Refetch the list of <li> elements
        li_elements = driver.find_elements(By.XPATH, "//*[@id='js-currency-sort-footer']//ul[@class='select-ul']//li")

        # Scroll the current element into view to make it interactable
        driver.execute_script("arguments[0].scrollIntoView(true);", li_elements[i])

        # Extract the currency text
        li_text = driver.execute_script("return arguments[0].querySelector('div.option > p').innerText;", li_elements[i])
        currency_symbol = li_text.split()[0]  # Extract the currency symbol (e.g., "$", "£", "€")
        print(f"Clicking on: {li_text.strip()}")

        # print(f'----{currency_symbol}----')

        # Click the currency option using JavaScript to avoid interactability issues
        driver.execute_script("arguments[0].click();", li_elements[i])

        # Wait for the page to update
        time.sleep(3)  # Adjust based on the site's responsiveness

        # Fetch updated prices after clicking the currency
        price_elements = driver.find_elements(By.CLASS_NAME, 'js-price-value')
        updated_prices = [elem.text for elem in price_elements]

        print(f"Found {len(updated_prices)} updated price elements.")

        default_price_element = driver.find_element(By.ID, "js-default-price")

        # Extract the text from the default price element
        default_price_text = default_price_element.text

        # Compare initial and updated prices
        for idx, (initial, updated) in enumerate(zip(initial_prices, updated_prices)):

            # Check if the clicked currency is the default one (EUR)
            if currency_symbol == default_currency:  # If the currency is the default (EUR) and clicked again, consider it a change
                price_changed = True
            else:
                price_changed = initial != updated

            result = {
                'Currency': li_text.strip(),
                'Cart Element': idx + 1,
                'Initial Price': initial,
                'Updated Price': updated,
                'Default Price': default_price_text,
                'Price Changed': price_changed
            }
            results.append(result)

finally:
    # Close the browser
    driver.quit()

    # Convert the results into a DataFrame
    df = pd.DataFrame(results)

    # Define the Excel file name
    excel_file = 'test_report.xlsx'

    # Try to load the workbook and append data to the existing sheet if it exists
    try:
        # Load the existing workbook
        workbook = load_workbook(excel_file)

        # Check if the sheet already exists
        if "Currency_Price_Comparison" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Currency_Price_Comparison")
        else:
            worksheet = workbook["Currency_Price_Comparison"]

        # If the sheet is empty, add headers
        if worksheet.max_row == 1:
            worksheet.append(["Currency", "Price Element", "Initial Price", "Updated Price", "Default Price", "Price Changed"])

        # Append results to the sheet
        for _, row in df.iterrows():
            worksheet.append(row.tolist())

        # Save the workbook
        workbook.save(excel_file)
        print(f"Results successfully appended to '{excel_file}'.")

    except Exception as e:
        print(f"Error occurred while appending to the Excel file: {str(e)}")
