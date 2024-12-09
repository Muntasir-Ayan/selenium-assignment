from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import pandas as pd
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time

# Function to extract data from the ScriptData object in the browser
def extract_data_from_script(driver):
    # Execute JavaScript to get the ScriptData from the console
    script = "return ScriptData;"  # Access the ScriptData object in the global window object
    script_data = driver.execute_script(script)  # Executes the script and returns the data
    
    if not script_data:
        print("No ScriptData found")
        return None
    
    # Extract the required fields from the ScriptData object
    data = {
        "SiteURL": script_data.get("config", {}).get("SiteUrl", "N/A"),
        "CampaignID": script_data.get("pageData", {}).get("CampaignId", "N/A"),
        "SiteName": script_data.get("config", {}).get("SiteName", ""),
        "Browser": driver.capabilities.get('browserName', 'N/A'),  # Browser info from WebDriver capabilities
        "CountryCode": script_data.get("userInfo", {}).get("CountryCode", ""),
        "IP": script_data.get("userInfo", {}).get("IP", "N/A")  # IP extraction if available
    }
    
    # Printing the extracted data for debugging purposes
    print(data)
    
    return data


# Function to save data to Excel (overwrite existing sheet or create a new one)
def save_to_excel(data, filename):
    try:
        # Check if the file exists
        try:
            # Try to load the existing workbook
            workbook = load_workbook(filename)
            print("Workbook loaded successfully.")
        except FileNotFoundError:
            # If file doesn't exist, create a new workbook
            workbook = Workbook()
            print("New workbook created.")

        # Check if the sheet exists
        if "Scraped Data" in workbook.sheetnames:
            worksheet = workbook["Scraped Data"]
            print("Sheet 'Scraped Data' found. Overwriting it.")
        else:
            # If the sheet does not exist, create it
            worksheet = workbook.create_sheet("Scraped Data")
            print("Sheet 'Scraped Data' created.")

        # If the sheet is empty, add headers
        if worksheet.max_row == 1:
            worksheet.append(["SiteURL", "CampaignID", "SiteName", "Browser", "CountryCode", "IP"])

        # Add the data to the sheet
        worksheet.append([data["SiteURL"], data["CampaignID"], data["SiteName"], data["Browser"], data["CountryCode"], data["IP"]])

        # Save the workbook
        workbook.save(filename)
        print("Data saved to Excel successfully!")
    except Exception as e:
        print(f"Error saving to Excel: {e}")


# Main function to scrape data
def main():
    # Set up the WebDriver (use ChromeDriverManager to manage the ChromeDriver)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    
    # URL to scrape
    url = "https://www.alojamiento.io/property/campo-lindo-apartment/BC-1935047"
    
    # Open the page
    driver.get(url)
    
    # Wait for the page to load (adjust sleep time as needed)
    time.sleep(5)

    # Extract the data from ScriptData
    data = extract_data_from_script(driver)
    
    if data:
        # Save the extracted data to an Excel file
        save_to_excel(data, 'test_report.xlsx')

    # Close the WebDriver
    driver.quit()


# Run the script
if __name__ == "__main__":
    main()
