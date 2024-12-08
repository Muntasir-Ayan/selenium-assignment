from selenium import webdriver
from selenium.webdriver.common.by import By
import pandas as pd
from openpyxl import load_workbook

# Function to test the HTML tag sequence (H1-H6)
def test_heading_sequence(url, testcase):
    result = {"page_url": url, "testcase": testcase, "passed/fail": "", "comments": ""}
    
    # Initialize the WebDriver
    driver = webdriver.Chrome()

    try:
        # Open the URL
        driver.get(url)

        # Extract all heading tags (h1 to h6)
        headings = {}
        for i in range(1, 7):
            headings[f"h{i}"] = driver.find_elements(By.TAG_NAME, f"h{i}")
        
        # Check if all heading tags are present and in sequence
        for i in range(1, 7):
            if len(headings[f"h{i}"]) == 0:  # If any heading tag is missing
                result["passed/fail"] = "fail"
                result["comments"] = f"Missing <h{i}> tag"
                return result

        # If all headings are present, check for the sequence (h1 should exist before h2, etc.)
        for i in range(1, 6):
            if len(headings[f"h{i}"]) > 0 and len(headings[f"h{i+1}"]) == 0: 
                result["passed/fail"] = "fail"
                result["comments"] = f"Sequence broken, <h{i+1}> tag missing after <h{i}>"
                return result

        # If all tags are in sequence
        result["passed/fail"] = "passed"
        result["comments"] = "All heading tags (h1-h6) are present and in correct sequence."

    except Exception as e:
        result["passed/fail"] = "fail"
        result["comments"] = str(e)
    
    finally:
        # Close the driver after use
        driver.quit()

    return result

# Function to add results to the Excel file
def add_to_excel(test_results, excel_file):
    try:
        # Load the existing workbook
        workbook = load_workbook(excel_file)
        
        # Check if the sheet already exists
        if "Heading_Sequence_Test" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Heading_Sequence_Test")
        else:
            worksheet = workbook["Heading_Sequence_Test"]
        
        # If the sheet is empty, add headers
        if worksheet.max_row == 1:
            worksheet.append(["page_url", "testcase", "passed/fail", "comments"])
        
        # Add each result to the sheet
        for result in test_results:
            worksheet.append([result["page_url"], result["testcase"], result["passed/fail"], result["comments"]])
        
        # Save the workbook with the new sheet data
        workbook.save(excel_file)
        print("Test results added to the 'Heading_Sequence_Test' sheet successfully!")

    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")

# List of test cases (URLs and test case names)
test_cases = [
    {"url": "https://www.alojamiento.io/property/campo-lindo-apartment/BC-1935047", "testcase": "Heading Sequence Test"},
    # Add more test cases here if needed
]

# List to hold test results
test_results = []

# Run the tests for each case
for case in test_cases:
    result = test_heading_sequence(case["url"], case["testcase"])
    test_results.append(result)

# Add the results to an existing Excel file
add_to_excel(test_results, "test_report.xlsx")


print("HTML sequence Test report generated successfully!")