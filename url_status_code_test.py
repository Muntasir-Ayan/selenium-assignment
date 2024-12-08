import requests
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import load_workbook

# Function to extract unique URLs from a page dynamically using Selenium
def get_urls_from_page(url):
    # Start the Chrome driver
    driver = webdriver.Chrome()
    driver.get(url)
    
    # Find all the <a> tags and extract the href attributes (URLs)
    links = driver.find_elements(By.TAG_NAME, 'a')
    
    # Use a set to store unique URLs (this filters out duplicates)
    urls = set(link.get_attribute('href') for link in links if link.get_attribute('href') is not None)
    
    driver.quit()
    return list(urls)  # Convert the set back to a list

# Function to test the status code of a URL
def test_url_status_code(url, testcase):
    result = {
        "page_url": url,
        "testcase": testcase,
        "status_code": None,
        "passed/fail": None,
        "comments": None
    }
    
    try:
        # Send a GET request to the URL
        response = requests.get(url)
        
        # Check the status code
        result["status_code"] = response.status_code
        if response.status_code == 404:
            result["passed/fail"] = "fail"
            result["comments"] = "URL returned 404"
        elif response.status_code == 403:
            result["passed/fail"] = "passed"
            result["comments"] = "Status code 403 (Forbidden)"
        elif response.status_code == 200:
            result["passed/fail"] = "passed"
            result["comments"] = "Status code 200 (OK)"
        else:
            result["passed/fail"] = "passed"
            result["comments"] = f"Status code {response.status_code}"
    
    except requests.exceptions.RequestException as e:
        result["passed/fail"] = "fail"
        result["comments"] = f"Error: {str(e)}"
    
    return result

# Function to add results to the Excel file
def add_to_excel(test_results, excel_file):
    try:
        # Load the existing workbook
        workbook = load_workbook(excel_file)
        
        # Check if the sheet already exists
        if "URL_Status_Code_Test" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("URL_Status_Code_Test")
        else:
            worksheet = workbook["URL_Status_Code_Test"]
        
        # If the sheet is empty, add headers
        if worksheet.max_row == 1:
            worksheet.append(["page_url", "testcase", "status_code", "passed/fail", "comments"])
        
        # Add each result to the sheet
        for result in test_results:
            worksheet.append([result["page_url"], result["testcase"], result["status_code"], result["passed/fail"], result["comments"]])
        
        # Save the workbook with the new sheet data
        workbook.save(excel_file)
        print("URL status code test results added to the 'URL_Status_Code_Test' sheet successfully!")

    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")

# Main function to execute the whole process
def main():
    # The URL from where we will extract the links
    page_url = "https://www.alojamiento.io/property/campo-lindo-apartment/BC-1935047"
    
    # First, get all the unique URLs from the page
    urls = get_urls_from_page(page_url)
    
    # List to hold test results
    test_results = []
    
    # Run the status code test for each extracted URL
    for url in urls:
        result = test_url_status_code(url, "URL Status Code Test")
        test_results.append(result)  # Add the result to the final results

    # Add the results to an existing Excel file
    add_to_excel(test_results, "test_report.xlsx")

    print("URL status code test report generated successfully!")

# Run the script
if __name__ == "__main__":
    main()
