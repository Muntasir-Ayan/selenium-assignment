from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
from openpyxl import load_workbook
import time

# Function to test the currency filter
def test_currency_filter(url, testcase):
    result = {"page_url": url, "testcase": testcase, "passed/fail": "", "comments": ""}
    
    # Initialize the WebDriver
    driver = webdriver.Chrome()

    try:
        # Open the URL
        driver.get(url)
        driver.maximize_window()
        time.sleep(2)  # Wait for the page to load

        # Locate the currency dropdown
        currency_dropdown = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "js-currency-sort-footer"))
        )

        # Scroll into view and open the dropdown
        driver.execute_script("arguments[0].scrollIntoView(true);", currency_dropdown)
        ActionChains(driver).move_to_element(currency_dropdown).click().perform()
        time.sleep(1)

        # Get all currency options
        currency_options = driver.find_elements(By.CSS_SELECTOR, "#js-currency-sort-footer .select-ul li")
        if not currency_options:
            result["passed/fail"] = "fail"
            result["comments"] = "No currency options found in the dropdown."
            return result

        # Test each currency
        for option in currency_options:
            currency_code = option.get_attribute("data-currency-country")
            currency_symbol = option.find_element(By.CSS_SELECTOR, "p").text.strip()
            
            # Click on the currency option
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", option)
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable(option)).click()
            except Exception as e:
                driver.execute_script("arguments[0].click();", option)  # Fallback to JavaScript click
            
            time.sleep(2)  # Wait for the page to update

            # Verify that all property tiles display the selected currency
            property_tiles = driver.find_elements(By.CSS_SELECTOR, ".property-tile-price")
            all_tiles_correct = all(currency_symbol in tile.text for tile in property_tiles)

            if not all_tiles_correct:
                result["passed/fail"] = "fail"
                result["comments"] = f"Currency {currency_symbol} mismatch found in property tiles."
                return result
        
        # If all currency options passed
        result["passed/fail"] = "passed"
        result["comments"] = "All currencies displayed correctly in property tiles."

    except Exception as e:
        result["passed/fail"] = "fail"
        result["comments"] = str(e)

    finally:
        # Close the driver after use
        driver.quit()

    return result

# Function to add results to the Excel file
def add_to_excel(test_results, excel_file):
    try:
        # Load the existing workbook
        workbook = load_workbook(excel_file)
        
        # Check if the sheet already exists
        if "Currency_Filter_Test" not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Currency_Filter_Test")
        else:
            worksheet = workbook["Currency_Filter_Test"]
        
        # If the sheet is empty, add headers
        if worksheet.max_row == 1:
            worksheet.append(["page_url", "testcase", "passed/fail", "comments"])
        
        # Add each result to the sheet
        for result in test_results:
            worksheet.append([result["page_url"], result["testcase"], result["passed/fail"], result["comments"]])
        
        # Save the workbook with the new sheet data
        workbook.save(excel_file)
        print("Test results added to the 'Currency_Filter_Test' sheet successfully!")

    except Exception as e:
        print(f"Error saving to Excel: {str(e)}")

# List of test cases (URLs and test case names)
test_cases = [
    {"url": "https://www.alojamiento.io/property/campo-lindo-apartment/BC-1935047", "testcase": "Currency Filter Test"},
    # Add more test cases here if needed
]

# List to hold test results
test_results = []

# Run the tests for each case
for case in test_cases:
    result = test_currency_filter(case["url"], case["testcase"])
    test_results.append(result)

# Add the results to an existing Excel file
add_to_excel(test_results, "test_report.xlsx")

print("Currency Filter Test report generated successfully!")
